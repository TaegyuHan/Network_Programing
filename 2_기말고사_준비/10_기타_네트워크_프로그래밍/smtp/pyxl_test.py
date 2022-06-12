"""
    메일 보내기: 단체 메일 전송하기
"""

from openpyxl import Workbook
import datetime

wb = Workbook()

ws = wb.active  # 활성화된 워크시트 선택
ws["A1"] = 42  # 셀에 데이터 추가
ws.append([1, 2, 3])  # 그다음 행에 데이터 추가
ws["A2"] = datetime.datetime.now()

wb.save("sample.xlsx")  # 저장